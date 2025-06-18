import os
import whisper
import pandas as pd
from langchain.chat_models import ChatOpenAI
from langchain.schema import HumanMessage
from config import OPENAI_API_KEY
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook


def append_to_excel(filename, row):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Filename", "Transcription", "Translation", "Segment_Timestamps", "Word_Timestamps"])
        wb.save(filename)
    wb = load_workbook(filename)
    ws = wb.active
    ws.append(row)
    wb.save(filename)


def transcribe_and_translate(file_path, filename, chat):
    try:
        model = whisper.load_model("whisper-turbo")
        transcript = model.transcribe(file_path, word_timestamps=True)
        jp_text = transcript["text"]
        segments = transcript["segments"]
        
        print(f"Transcription of {filename} complete.")
        
        segment_timestamps = []
        word_timestamps = []
        
        for segment in segments:
            segment_start = segment.get("start", 0)
            segment_end = segment.get("end", 0)
            segment_text = segment.get("text", "")
            segment_timestamps.append(f"[{segment_start:.2f}-{segment_end:.2f}] {segment_text.strip()}")
            
            if "words" in segment:
                words_with_times = []
                for word_info in segment["words"]:
                    word_text = word_info.get("word", "")
                    word_start = word_info.get("start", 0)
                    word_end = word_info.get("end", 0)
                    words_with_times.append(f"{word_text}[{word_start:.2f}-{word_end:.2f}]")
                word_timestamps.append(" ".join(words_with_times))
            else:
                word_timestamps.append("")
        
        try:
            prompt = f"Translate the following Japanese text to English.\nJapanese: {jp_text}\nEnglish:"
            response = chat([HumanMessage(content=prompt)])
            en_text = response.content.strip()
        except Exception as e:
            print(f"Error translating {filename}: {e}")
            en_text = f"Error: {e}"
        
        segment_times_str = " | ".join(segment_timestamps)
        word_times_str = " | ".join(word_timestamps)
        
        return [filename, jp_text, en_text, segment_times_str, word_times_str]
    except Exception as e:
        print(f"Error transcribing {filename}: {e}")
        return [filename, f"Error: {e}", "", "", ""]


def transcribe_folder(folder_path, output_excel_file):
    """
    Transcribes all audio files in a folder using Whisper Turbo with timestamps, translates the results from Japanese to English using ChatOpenAI, and saves the results to an Excel file incrementally.
    """
    audio_files = [f for f in os.listdir(folder_path) if f.endswith((".mp3", ".wav", ".m4a", ".aac", ".flac"))]
    if not audio_files:
        print(f"No audio files found in the folder: {folder_path}")
        return

    os.environ["OPENAI_API_KEY"] = OPENAI_API_KEY
    chat = ChatOpenAI(temperature=0, model_name="gpt-3.5-turbo")

    with ThreadPoolExecutor(max_workers=2) as executor:
        futures = []
        for filename in audio_files:
            file_path = os.path.join(folder_path, filename)
            print(f"Submitting transcription and translation for: {filename}")
            future = executor.submit(transcribe_and_translate, file_path, filename, chat)
            futures.append(future)
        for future in as_completed(futures):
            row = future.result()
            append_to_excel(output_excel_file, row)
            print(f"Result for {row[0]} written to {output_excel_file}")

    print(f"All transcriptions and translations saved to: {output_excel_file}")

if __name__ == "__main__":
    folder_path = "audio/"  # Default path to the audio folder
    output_excel_file = "transcriptions.xlsx"  # Default output Excel file name
    transcribe_folder(folder_path, output_excel_file)
